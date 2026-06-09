# Copyright (c) 2026, Imogi and contributors

import io

import frappe
from frappe import _
from frappe.utils import cint, flt
from frappe.utils.xlsxutils import read_xls_file_from_attached_file

from imogi_pos.imogi_pos.utils.import_helpers import (
	ensure_uom,
	get_buying_price_list,
	resolve_item_code,
	slug_item_code,
	upsert_buying_item_price,
)
from imogi_pos.imogi_pos.utils.import_helpers import ensure_import_settings
from imogi_pos.imogi_pos.utils.menu_import_helpers import (
	_get_file_doc,
	_file_extension,
	_load_csv_rows,
	require_import_access,
	sheet_rows_to_dicts,
)

STOCK_ITEM_KEYS = (
	"komponen",
	"bom_product",
	"component",
	"bahan",
	"bahan baku",
	"bahan_baku",
	"item",
	"item_code",
	"item_name",
	"nama item",
	"nama_item",
	"produk",
)
STOCK_QTY_KEYS = ("qty", "quantity", "jumlah", "stock", "stok")
STOCK_UOM_KEYS = ("uom", "satuan", "unit")
STOCK_RATE_KEYS = ("harga", "rate", "price", "biaya", "valuation_rate", "harga satuan", "harga_satuan")

STOCK_SHEET_NAMES = (
	"stok awal",
	"stok",
	"stock",
	"stok bahan",
	"stok_bahan",
	"inventory",
	"opening stock",
)


def _normalize_stock_row(raw):
	out = {}
	for key, value in raw.items():
		normalized_key = str(key or "").strip().lower()
		if value is None:
			out[normalized_key] = ""
		elif isinstance(value, str):
			out[normalized_key] = value.strip()
		elif isinstance(value, float) and value == int(value):
			out[normalized_key] = str(int(value))
		else:
			out[normalized_key] = str(value).strip()
	if raw.get("_row") is not None:
		out["_row"] = raw.get("_row")
	return out


def _pick(row, keys):
	for key in keys:
		value = row.get(key)
		if value not in (None, ""):
			return str(value).strip()
	return ""


def parse_stock_rows(rows):
	parsed = []
	for raw in rows:
		row = _normalize_stock_row(raw)
		item_ref = _pick(row, STOCK_ITEM_KEYS)
		qty = flt(_pick(row, STOCK_QTY_KEYS))
		if not item_ref or qty <= 0:
			continue
		parsed.append(
			{
				"item_ref": item_ref,
				"qty": qty,
				"uom": _pick(row, STOCK_UOM_KEYS),
				"rate": flt(_pick(row, STOCK_RATE_KEYS)),
				"_row": row.get("_row"),
			}
		)
	return parsed


def read_stock_sheet_rows(content, ext):
	if ext == "xlsx":
		from openpyxl import load_workbook

		wb = load_workbook(filename=io.BytesIO(content), data_only=True)
		ws = None
		for sheet_name in wb.sheetnames:
			if sheet_name.strip().lower() in STOCK_SHEET_NAMES:
				ws = wb[sheet_name]
				break
		if ws is None:
			return None

		rows = []
		for row in ws.iter_rows():
			rows.append([cell.value for cell in row])
		return rows

	if ext == "xls":
		# Legacy xls: only one sheet — treat as stock if headers match.
		return read_xls_file_from_attached_file(content)

	return None


def load_stock_rows_from_file(file_url=None, csv_text=None, optional=False):
	if file_url and not csv_text:
		file_doc = _get_file_doc(file_url)
		ext = _file_extension(file_doc)
		content = file_doc.get_content()

		if ext in ("xlsx", "xls"):
			raw_rows = read_stock_sheet_rows(content, ext)
			if not raw_rows:
				if optional:
					return []
				frappe.throw(_("Sheet 'Stok Awal' tidak ditemukan di file Excel"))
			rows = sheet_rows_to_dicts(raw_rows)
			if ext == "xls" and optional:
				# Avoid mis-reading a menu xls as stock when optional.
				parsed = parse_stock_rows(rows)
				if not parsed:
					return []
			return rows

		if isinstance(content, bytes):
			csv_text = content.decode("utf-8-sig")
		else:
			csv_text = content or ""

	if not csv_text:
		if optional:
			return []
		frappe.throw(_("Upload file Excel/CSV stok"))

	return _load_csv_rows(csv_text)


def _resolve_stock_item_code(item_ref):
	return resolve_item_code(item_ref) or slug_item_code(item_ref)


def _update_item_valuation(item_code, rate):
	if rate <= 0:
		return
	frappe.db.set_value("Item", item_code, "valuation_rate", rate)


def create_material_receipt(company, warehouse, lines):
	if not warehouse:
		frappe.throw(_("Set Gudang Default di IMOGI POS Settings"))

	cost_center = frappe.get_cached_value("Company", company, "cost_center")
	expense_account = frappe.get_cached_value("Company", company, "stock_adjustment_account")

	se = frappe.new_doc("Stock Entry")
	se.stock_entry_type = "Material Receipt"
	se.purpose = "Material Receipt"
	se.company = company

	for line in lines:
		item_code = line["item_code"]
		qty = flt(line["qty"])
		if qty <= 0:
			continue

		uom = ensure_uom(line.get("uom") or frappe.db.get_value("Item", item_code, "stock_uom") or "Nos")
		rate = flt(line.get("rate") or 0)
		if not rate:
			rate = flt(frappe.db.get_value("Item", item_code, "valuation_rate"))

		row = {
			"item_code": item_code,
			"t_warehouse": warehouse,
			"qty": qty,
			"uom": uom,
			"stock_uom": uom,
			"conversion_factor": 1,
			"transfer_qty": qty,
			"basic_rate": rate,
		}
		if cost_center:
			row["cost_center"] = cost_center
		if expense_account:
			row["expense_account"] = expense_account
		se.append("items", row)

	if not se.items:
		return None

	se.set_stock_entry_type()
	se.insert(ignore_permissions=True)
	se.submit()
	return se.name


def run_stock_import(rows, settings=None, update_rate=1, warehouse=None):
	settings = settings or ensure_import_settings()
	if not settings.default_company:
		frappe.throw(_("Set Perusahaan di IMOGI POS Settings"))
	warehouse = warehouse or settings.default_warehouse
	if not warehouse:
		frappe.throw(_("Set Gudang Default di IMOGI POS Settings"))

	parsed = parse_stock_rows(rows)
	if not parsed:
		frappe.throw(_("File tidak memiliki baris stok yang valid"))

	stats = {
		"stock_imported": True,
		"stock_lines": 0,
		"stock_skipped": 0,
		"buying_prices": 0,
		"stock_entry": None,
		"errors": [],
	}

	buying_price_list = get_buying_price_list(settings.default_company) if cint(update_rate) else None
	buying_price_list_warned = False

	receipt_lines = []
	for line in parsed:
		item_ref = line["item_ref"]
		item_code = _resolve_stock_item_code(item_ref)
		if not frappe.db.exists("Item", item_code):
			stats["errors"].append(
				{
					"row": line.get("_row"),
					"item_code": item_ref,
					"error": _("Item {0} belum ada. Import menu/BOM terlebih dahulu.").format(item_ref),
				}
			)
			stats["stock_skipped"] += 1
			continue

		if cint(update_rate) and line.get("rate"):
			_update_item_valuation(item_code, line["rate"])
			if buying_price_list:
				if upsert_buying_item_price(
					item_code,
					line["rate"],
					uom=line.get("uom"),
					company=settings.default_company,
				):
					stats["buying_prices"] += 1
			elif not buying_price_list_warned:
				stats["errors"].append(
					{
						"row": None,
						"item_code": None,
						"error": _(
							"Buying Price List belum di-set di Buying Settings — harga beli tidak diimport."
						),
					}
				)
				buying_price_list_warned = True

		rate = flt(line.get("rate") or frappe.db.get_value("Item", item_code, "valuation_rate"))
		receipt_lines.append(
			{
				"item_code": item_code,
				"qty": line["qty"],
				"uom": line.get("uom"),
				"rate": rate,
			}
		)
		stats["stock_lines"] += 1

	if not receipt_lines:
		stats["stock_imported"] = False
		return stats

	try:
		stats["stock_entry"] = create_material_receipt(
			settings.default_company,
			warehouse,
			receipt_lines,
		)
	except Exception as exc:
		stats["stock_imported"] = False
		stats["errors"].append({"row": None, "item_code": None, "error": str(exc)})

	return stats


def try_import_stock_from_file(file_url, settings, update_rate=1, warehouse=None):
	rows = load_stock_rows_from_file(file_url=file_url, optional=True)
	if not rows:
		return {
			"stock_imported": False,
			"stock_lines": 0,
			"stock_skipped": 0,
			"stock_entry": None,
			"errors": [],
		}

	parsed = parse_stock_rows(rows)
	if not parsed:
		return {
			"stock_imported": False,
			"stock_lines": 0,
			"stock_skipped": 0,
			"stock_entry": None,
			"errors": [],
		}

	return run_stock_import(rows, settings, update_rate=cint(update_rate), warehouse=warehouse)
