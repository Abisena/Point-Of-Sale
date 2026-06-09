# Copyright (c) 2026, Imogi and contributors

import csv
import io
import re

import frappe
from frappe import _
from frappe.utils import cint, flt
from frappe.utils.xlsxutils import read_xls_file_from_attached_file

from imogi_pos.imogi_pos.utils.import_helpers import (
	ensure_import_settings,
	ensure_item_group,
	ensure_uom,
	normalize_category,
	resolve_item_code,
	slug_item_code,
)

PRODUCT_KEYS = (
	"produk",
	"product",
	"name",
	"nama item",
	"nama_item",
	"nama produk",
	"nama_produk",
	"item_name",
	"finished product",
	"finished_product",
)
COMPONENT_KEYS = (
	"komponen",
	"bom_product",
	"component",
	"bahan",
	"bahan baku",
	"bahan_baku",
	"nama komponen",
	"nama_komponen",
	"raw material",
	"raw_material",
	"materials",
	"material",
)
CATEGORY_KEYS = ("kategori", "category", "imogi_pos_category")
SELLING_PRICE_KEYS = (
	"harga jual",
	"harga_jual",
	"standard_rate",
	"selling_price",
	"harga jual pos",
	"harga_jual_pos",
)
COST_KEYS = ("harga", "price", "biaya", "cost", "rate")
QTY_KEYS = ("qty", "quantity", "jumlah")
UOM_KEYS = ("uom", "satuan", "unit")
ADDON_KEYS = ("add_on", "add_ons", "addon")

IMOGI_HEADER_KEYS = frozenset(
	{
		"produk",
		"product",
		"komponen",
		"component",
		"qty",
		"uom",
		"harga",
		"temperatur",
		"ukuran",
		"warna",
		"kategori",
	}
)

QTY_UOM_RE = re.compile(r"^([\d.,]+)\s+(.+)$", re.IGNORECASE)


def require_import_access():
	if frappe.session.user == "Guest":
		frappe.throw(_("Login required"), frappe.AuthenticationError)
	if not frappe.has_permission("Item", "create"):
		frappe.throw(_("Not permitted to import products"), frappe.PermissionError)


def _get_file_doc(file_url):
	file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not file_name:
		frappe.throw(_("Uploaded file not found"))
	return frappe.get_doc("File", file_name)


def _file_extension(file_doc):
	name = (file_doc.file_name or file_doc.file_url or "").lower()
	if name.endswith(".xlsx"):
		return "xlsx"
	if name.endswith(".xls"):
		return "xls"
	return "csv"


def _cell_to_text(value):
	if value is None:
		return ""
	if isinstance(value, float) and value == int(value):
		return str(int(value))
	return str(value).strip()


def sheet_rows_to_dicts(rows):
	if not rows:
		frappe.throw(_("File tidak memiliki data"))

	headers = [str(h or "").strip() for h in rows[0]]
	if not any(headers):
		frappe.throw(_("File tidak memiliki baris header"))

	out = []
	for row_idx, row in enumerate(rows[1:], start=2):
		if not any(v not in (None, "") for v in row):
			continue
		record = {"_row": row_idx}
		for i, header in enumerate(headers):
			if not header:
				continue
			value = row[i] if i < len(row) else None
			record[header] = _cell_to_text(value)
		out.append(record)
	return out


def _normalize_header_key(value):
	return str(value or "").strip().lower()


def _rows_have_imogi_headers(headers):
	normalized = {_normalize_header_key(header) for header in headers if _normalize_header_key(header)}
	return bool(normalized & {"produk", "product"} or normalized & {"komponen", "component"})


def _parse_legacy_cost(value):
	text = (value or "").strip().replace(" ", "")
	if not text:
		return ""
	if "," in text and "." in text:
		return text.replace(",", "")
	if "," in text:
		parts = text.split(",")
		if len(parts) == 2 and len(parts[1]) <= 2:
			return f"{parts[0].replace('.', '')}.{parts[1]}"
		return text.replace(",", "")
	return text


def _split_qty_uom(value):
	text = (value or "").strip()
	if not text:
		return "", ""
	match = QTY_UOM_RE.match(text)
	if match:
		return match.group(1).replace(",", ""), match.group(2).strip()
	return text, ""


def _legacy_product_record(product_name, row_idx, category="Beverage"):
	from imogi_pos.imogi_pos.utils.menu_variant_import import resolve_product_variant

	base_name, attrs = resolve_product_variant(product_name, {})
	return {
		"Produk": base_name,
		"Temperatur": attrs.get("Temperatur", ""),
		"Ukuran": attrs.get("Ukuran", ""),
		"Warna": attrs.get("Warna", ""),
		"Komponen": "",
		"Qty": "",
		"UOM": "",
		"Harga": "",
		"Kategori": category,
		"Harga Jual": "",
		"_row": row_idx,
	}


def _legacy_component_record(component, qty_uom, cost, row_idx):
	qty, uom = _split_qty_uom(qty_uom)
	return {
		"Produk": "",
		"Komponen": component,
		"Qty": qty,
		"UOM": uom,
		"Harga": _parse_legacy_cost(cost),
		"_row": row_idx,
	}


def convert_legacy_bom_rows(raw_rows):
	"""Convert headerless Dirno BOM export (4 kolom) to IMOGI import rows."""
	rows_out = []

	for row_idx, raw in enumerate(raw_rows, start=1):
		cells = [_cell_to_text(cell) for cell in raw]
		while len(cells) < 4:
			cells.append("")

		col_a, col_b, col_c, col_d = cells[0], cells[1], cells[2], cells[3]
		if not any(cells):
			continue

		if row_idx == 1 and _normalize_header_key(col_a) in ("produk", "product"):
			continue

		if col_a and col_b:
			rows_out.append(_legacy_product_record(col_a, row_idx))
			rows_out.append(_legacy_component_record(col_b, col_c, col_d, row_idx))
			continue

		if col_a and not col_b:
			rows_out.append(_legacy_product_record(col_a, row_idx))
			continue

		if col_b:
			rows_out.append(_legacy_component_record(col_b, col_c, col_d, row_idx))

	return rows_out


def _looks_like_legacy_bom(raw_rows):
	if not raw_rows:
		return False
	headers = [str(cell or "") for cell in raw_rows[0]]
	if _rows_have_imogi_headers(headers):
		return False

	for raw in raw_rows[:40]:
		cells = [_cell_to_text(cell) for cell in raw]
		if len(cells) < 2:
			continue
		col_b = (cells[1] or "").strip()
		if not col_b:
			continue
		if _normalize_header_key(col_b) in IMOGI_HEADER_KEYS:
			continue
		return True
	return False


def normalize_menu_sheet_rows(raw_rows):
	if not raw_rows:
		frappe.throw(_("File tidak memiliki data"))
	if _looks_like_legacy_bom(raw_rows):
		return convert_legacy_bom_rows(raw_rows)
	return sheet_rows_to_dicts(raw_rows)


def _load_csv_rows(csv_text):
	if not csv_text:
		frappe.throw(_("Upload file CSV/Excel atau tempel teks CSV"))

	if not csv_text.strip():
		frappe.throw(_("Upload file CSV/Excel atau tempel teks CSV"))

	reader = csv.reader(io.StringIO(csv_text))
	raw_rows = list(reader)
	if not raw_rows:
		frappe.throw(_("File tidak memiliki baris header"))

	if _looks_like_legacy_bom(raw_rows):
		return convert_legacy_bom_rows(raw_rows)

	dict_reader = csv.DictReader(io.StringIO(csv_text))
	if not dict_reader.fieldnames:
		frappe.throw(_("File tidak memiliki baris header"))

	rows = []
	for idx, raw in enumerate(dict_reader, start=2):
		row = dict(raw)
		row["_row"] = idx
		rows.append(row)
	return rows


def load_menu_rows_from_file(file_url=None, csv_text=None):
	if file_url and not csv_text:
		file_doc = _get_file_doc(file_url)
		ext = _file_extension(file_doc)
		content = file_doc.get_content()

		if ext in ("xlsx", "xls"):
			raw_rows = read_menu_sheet_rows(content, ext)
			if not raw_rows:
				frappe.throw(_("File Excel tidak bisa dibaca"))
			return normalize_menu_sheet_rows(raw_rows)

		if isinstance(content, bytes):
			csv_text = content.decode("utf-8-sig")
		else:
			csv_text = content or ""

	return _load_csv_rows(csv_text)


def _pick(row, keys):
	for key in keys:
		value = row.get(key)
		if value not in (None, ""):
			return str(value).strip()
	return ""


def _pick_number(row, keys):
	value = _pick(row, keys)
	return flt(value) if value else 0


def _normalize_menu_row(raw):
	out = {}
	for key, value in raw.items():
		normalized_key = str(key or "").strip().lower()
		if value is None:
			out[normalized_key] = ""
		elif isinstance(value, str):
			out[normalized_key] = value.strip()
		elif isinstance(value, float) and value == int(value):
			out[normalized_key] = str(int(value))
		else:
			out[normalized_key] = str(value).strip()
	if raw.get("_row") is not None:
		out["_row"] = raw.get("_row")
	return out


def _is_summary_component(name):
	text = (name or "").strip().lower()
	return text in ("total", "subtotal", "jumlah", "grand total")


def _component_dict(row):
	return {
		"name": _pick(row, COMPONENT_KEYS),
		"qty": _pick(row, QTY_KEYS),
		"uom": _pick(row, UOM_KEYS),
		"harga": _pick(row, COST_KEYS),
		"_row": row.get("_row"),
	}


def _product_meta(row):
	return {
		"category": _pick(row, CATEGORY_KEYS),
		"selling_price": _pick(row, SELLING_PRICE_KEYS),
		"add_on": _pick(row, ADDON_KEYS),
		"row": row.get("_row"),
	}


def _fill_forward_menu_rows(rows):
	"""Carry product + variant columns down (Excel merged cells leave blanks)."""
	from imogi_pos.imogi_pos.utils.menu_variant_import import VARIANT_COLUMN_MAP

	ctx_product = ""
	ctx_variants = {}
	ctx_category = ""
	ctx_selling_price = ""

	for row in rows:
		product = _pick(row, PRODUCT_KEYS)
		if product:
			if product != ctx_product:
				ctx_variants = {}
			ctx_product = product

		for key in VARIANT_COLUMN_MAP:
			value = (row.get(key) or "").strip()
			if value:
				ctx_variants[key] = value

		category = _pick(row, CATEGORY_KEYS)
		if category:
			ctx_category = category

		selling_price = _pick(row, SELLING_PRICE_KEYS)
		if selling_price:
			ctx_selling_price = selling_price

		if not product and ctx_product:
			row["produk"] = ctx_product
			row["product"] = ctx_product
			row["name"] = ctx_product

		for key, value in ctx_variants.items():
			if not (row.get(key) or "").strip():
				row[key] = value

		if not _pick(row, CATEGORY_KEYS) and ctx_category:
			row["kategori"] = ctx_category
		if not _pick(row, SELLING_PRICE_KEYS) and ctx_selling_price:
			row["harga jual"] = ctx_selling_price

	return rows


def _flat_block_key(row):
	from imogi_pos.imogi_pos.utils.menu_variant_import import resolve_product_variant

	product = _pick(row, PRODUCT_KEYS)
	base_name, attrs = resolve_product_variant(product, row)
	parts = [base_name or product]
	for attribute in ("Temperatur", "Ukuran", "Warna"):
		value = attrs.get(attribute)
		if value:
			parts.append(f"{attribute}={value}")
	return "|".join(parts)


def parse_menu_blocks(rows):
	"""Parse flat rows into finished-product blocks with BOM components."""
	normalized = [_normalize_menu_row(row) for row in rows]
	if not normalized:
		return []

	normalized = _fill_forward_menu_rows(normalized)

	has_block_headers = any(_pick(row, PRODUCT_KEYS) and not _pick(row, COMPONENT_KEYS) for row in normalized)
	if has_block_headers:
		return _parse_block_format(normalized)
	return _parse_flat_format(normalized)


def _parse_block_format(rows):
	blocks = []
	current = None

	for row in rows:
		product = _pick(row, PRODUCT_KEYS)
		component = _pick(row, COMPONENT_KEYS)

		if product and not component:
			current = {
				"product_name": product,
				"components": [],
				**_product_meta(row),
			}
			from imogi_pos.imogi_pos.utils.menu_variant_import import enrich_block_variant_info

			enrich_block_variant_info(current, row)
			blocks.append(current)
			continue

		if component:
			if _is_summary_component(component):
				continue
			if not current:
				frappe.throw(
					_("Baris {0}: komponen '{1}' tanpa produk jadi di atasnya").format(
						row.get("_row") or "?", component
					)
				)
			current["components"].append(_component_dict(row))

	return [block for block in blocks if block.get("product_name")]


def _parse_flat_format(rows):
	grouped = {}

	for row in rows:
		product = _pick(row, PRODUCT_KEYS)
		component = _pick(row, COMPONENT_KEYS)
		if not product or not component or _is_summary_component(component):
			continue

		block_key = _flat_block_key(row)
		block = grouped.get(block_key)
		if not block:
			block = {
				"product_name": product,
				"components": [],
				**_product_meta(row),
			}
			from imogi_pos.imogi_pos.utils.menu_variant_import import enrich_block_variant_info

			enrich_block_variant_info(block, row)
			grouped[block_key] = block

		if not block.get("category"):
			block["category"] = _pick(row, CATEGORY_KEYS)
		if not block.get("selling_price"):
			block["selling_price"] = _pick(row, SELLING_PRICE_KEYS)
		if not block.get("add_on"):
			block["add_on"] = _pick(row, ADDON_KEYS)

		block["components"].append(_component_dict(row))

	return list(grouped.values())


def _line_unit_rate(harga, qty):
	cost = flt(harga)
	quantity = flt(qty)
	if quantity > 0 and cost > 0:
		return cost / quantity
	return cost


def upsert_component_from_menu(name, uom, harga, qty, update_existing=1):
	item_code = resolve_item_code(name) or slug_item_code(name)
	uom = ensure_uom(uom or "Nos")
	rate = _line_unit_rate(harga, qty)
	ensure_item_group("Bahan Baku")

	if frappe.db.exists("Item", item_code):
		if not cint(update_existing):
			return "skipped", item_code
		item = frappe.get_doc("Item", item_code)
		action = "updated"
	else:
		item = frappe.new_doc("Item")
		item.item_code = item_code
		item.is_stock_item = 1
		item.is_sales_item = 0
		item.is_purchase_item = 1
		action = "created"

	item.item_name = name
	item.item_group = "Bahan Baku"
	item.stock_uom = uom
	if rate > 0:
		item.valuation_rate = rate

	if action == "created":
		item.insert(ignore_permissions=True)
	else:
		item.save(ignore_permissions=True)

	return action, item_code


def _components_to_bom_lines(components):
	lines = []
	for component in components:
		name = component.get("name")
		if not name:
			continue
		qty = flt(component.get("qty"))
		if qty <= 0:
			continue
		lines.append(
			{
				"bom_product": name,
				"qty": qty,
				"uom": component.get("uom"),
				"rate": _line_unit_rate(component.get("harga"), qty),
				"_row": component.get("_row"),
			}
		)
	return lines


def _is_service_block(block):
	category = normalize_category(block.get("category") or "")
	return category.lower() == "service"


def run_menu_import(rows, settings=None, update_existing=1):
	settings = settings or ensure_import_settings()
	if not settings.default_company:
		frappe.throw(_("Set Perusahaan di IMOGI POS Settings"))

	blocks = parse_menu_blocks(rows)
	if not blocks:
		frappe.throw(_("File tidak memiliki data menu yang valid"))

	from imogi_pos.imogi_pos.utils.menu_variant_import import (
		import_standalone_block,
		import_variant_group,
		partition_blocks,
	)

	stats = {
		"products_created": 0,
		"products_updated": 0,
		"templates_created": 0,
		"templates_updated": 0,
		"variants_created": 0,
		"variants_updated": 0,
		"templates_demoted": 0,
		"variants_retired": 0,
		"variants_deleted": 0,
		"variants_reenabled": 0,
		"components_created": 0,
		"components_updated": 0,
		"boms_created": 0,
		"boms_updated": 0,
		"skipped_service": 0,
		"skipped": 0,
		"errors": [],
	}

	standalone_blocks, variant_groups = partition_blocks(blocks)

	for block in standalone_blocks:
		import_standalone_block(block, settings, update_existing, stats)

	for base_name, group_blocks in variant_groups.items():
		import_variant_group(base_name, group_blocks, settings, update_existing, stats)

	return stats


def read_menu_sheet_rows(content, ext):
	from frappe.utils.xlsxutils import read_xls_file_from_attached_file, read_xlsx_file_from_attached_file

	if ext == "xlsx":
		from openpyxl import load_workbook

		wb = load_workbook(filename=io.BytesIO(content), data_only=True)
		ws = None
		for sheet_name in wb.sheetnames:
			if sheet_name.strip().lower() in ("menu", "produk", "product", "products"):
				ws = wb[sheet_name]
				break
		if ws is None:
			ws = wb.active

		rows = []
		for row in ws.iter_rows():
			rows.append([cell.value for cell in row])
		return rows

	if ext == "xls":
		return read_xls_file_from_attached_file(content)

	return None
