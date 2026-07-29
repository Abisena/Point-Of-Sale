# Copyright (c) 2026, Imogi and contributors
"""Parsing mutasi rekening BCA (CSV/XLSX mentah, apa adanya dari e-banking)."""

import csv
import hashlib
import io
import re
from datetime import datetime

import frappe
from frappe import _
from frappe.utils import flt

from imogi_pos.imogi_pos.utils.menu_import_helpers import _file_extension, _get_file_doc

DATE_FORMATS = ("%d/%m/%Y", "%d/%m/%y")
AMOUNT_RE = re.compile(r"^([\d.,]+)\s*(CR|DB|DR)?$", re.IGNORECASE)


def require_bank_import_access():
	if frappe.session.user == "Guest":
		frappe.throw(_("Login required"), frappe.AuthenticationError)
	if not frappe.has_permission("Bank Transaction", "create"):
		frappe.throw(_("Tidak punya akses untuk mengimpor mutasi bank"), frappe.PermissionError)


def _read_excel_rows(content, ext):
	if ext == "xlsx":
		from openpyxl import load_workbook

		wb = load_workbook(filename=io.BytesIO(content), data_only=True)
		ws = wb.active
		return [[cell.value for cell in row] for row in ws.iter_rows()]

	from frappe.utils.xlsxutils import read_xls_file_from_attached_file

	return read_xls_file_from_attached_file(content)


def _detect_csv_delimiter(text):
	sample = text[:2048]
	try:
		return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
	except csv.Error:
		pass
	first = next((line for line in sample.splitlines() if line.strip()), "")
	for delim in ("\t", ";", ",", "|"):
		if first.count(delim) >= 1:
			return delim
	return ","


def _read_raw_rows(file_doc, ext):
	content = file_doc.get_content()
	if ext in ("xlsx", "xls"):
		rows = _read_excel_rows(content, ext)
		if not rows:
			frappe.throw(_("File Excel tidak bisa dibaca"))
		return rows

	text = content.decode("utf-8-sig") if isinstance(content, bytes) else (content or "")
	if not text.strip():
		frappe.throw(_("Upload file mutasi bank (.csv/.xlsx)"))
	delimiter = _detect_csv_delimiter(text)
	rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
	if not rows:
		frappe.throw(_("File tidak memiliki data"))
	return rows


def _locate_header_row(rows):
	for idx, row in enumerate(rows):
		cells = [str(c or "").strip().lower() for c in row]
		has_tanggal = any("tanggal" in c for c in cells)
		has_keterangan = any("keterangan" in c for c in cells)
		if has_tanggal and has_keterangan:
			return idx
	frappe.throw(
		_(
			"Tidak menemukan baris header transaksi (kolom Tanggal Transaksi/Keterangan) di file ini"
		)
	)


def _map_columns(header_row):
	cells = [str(c or "").strip().lower() for c in header_row]

	def _find(*tokens):
		for i, cell in enumerate(cells):
			if any(token in cell for token in tokens):
				return i
		return None

	return {
		"date": _find("tanggal"),
		"description": _find("keterangan"),
		"branch": _find("cabang"),
		"amount": _find("jumlah", "mutasi"),
		"balance": _find("saldo"),
	}


def _parse_bca_date(value):
	text = str(value or "").strip()
	if not text:
		return None
	for fmt in DATE_FORMATS:
		try:
			return datetime.strptime(text, fmt).date()
		except ValueError:
			continue
	return None


def _parse_amount(value):
	"""'187,000.00 CR' -> (deposit, withdrawal) = (187000.0, 0.0); '25,000.00 DB' -> (0.0, 25000.0)."""
	text = str(value or "").strip()
	if not text:
		return None
	match = AMOUNT_RE.match(text)
	if not match:
		return None
	amount = flt(match.group(1).replace(",", ""))
	kind = (match.group(2) or "").upper()
	if kind == "CR":
		return amount, 0.0
	if kind in ("DB", "DR"):
		return 0.0, amount
	return None


def _cell(row, idx):
	if idx is None or idx >= len(row):
		return None
	return row[idx]


def parse_bca_statement(file_url):
	"""Return (rows, errors). rows: list of dict(row, date, description, branch, deposit, withdrawal, balance)."""
	file_doc = _get_file_doc(file_url)
	ext = _file_extension(file_doc)
	raw_rows = _read_raw_rows(file_doc, ext)

	header_idx = _locate_header_row(raw_rows)
	columns = _map_columns(raw_rows[header_idx])

	parsed = []
	errors = []
	for row_idx, row in enumerate(raw_rows[header_idx + 1 :], start=header_idx + 2):
		if not any(c not in (None, "") for c in row):
			continue

		date = _parse_bca_date(_cell(row, columns["date"]))
		if not date:
			# baris footer/ringkasan (mis. "Saldo Awal"/"Bersambung...") tanpa tanggal transaksi valid
			continue

		amount_value = _cell(row, columns["amount"])
		amounts = _parse_amount(amount_value)
		if amounts is None:
			errors.append(
				{"row": row_idx, "error": _("Jumlah tidak bisa dibaca: {0}").format(amount_value)}
			)
			continue
		deposit, withdrawal = amounts

		description = str(_cell(row, columns["description"]) or "").strip()
		branch = str(_cell(row, columns["branch"]) or "").strip()
		balance = flt(str(_cell(row, columns["balance"]) or "0").replace(",", ""))

		parsed.append(
			{
				"row": row_idx,
				"date": date,
				"description": description,
				"branch": branch,
				"deposit": deposit,
				"withdrawal": withdrawal,
				"balance": balance,
			}
		)

	return parsed, errors


def row_transaction_id(bank_account, row):
	raw = "|".join(
		[
			bank_account,
			str(row["date"]),
			f"{row['deposit']:.2f}",
			f"{row['withdrawal']:.2f}",
			row["description"],
			f"{row['balance']:.2f}",
		]
	)
	return "BCA-" + hashlib.sha1(raw.encode("utf-8")).hexdigest()[:20]
