#!/usr/bin/env python3
"""Build Stok Awal CSV/XLSX from dirno_menu_import.csv BOM components."""

from __future__ import annotations

import csv
import io
from collections import OrderedDict
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
MENU_CSV = SCRIPT_DIR / "dirno_menu_import.csv"
STOCK_CSV = SCRIPT_DIR / "dirno_stok_awal.csv"
WORKBOOK_XLSX = SCRIPT_DIR / "dirno_menu_lengkap.xlsx"

STOCK_HEADERS = ["Komponen", "Qty", "UOM", "Harga"]
MENU_HEADERS = [
	"Produk",
	"Temperatur",
	"Ukuran",
	"Warna",
	"Komponen",
	"Qty",
	"UOM",
	"Harga",
	"Kategori",
	"Harga Jual",
]


def _parse_float(value):
	text = str(value or "").strip().replace(",", "")
	try:
		return float(text)
	except ValueError:
		return 0.0


def _unit_rate(qty, harga):
	qty = _parse_float(qty)
	harga = _parse_float(harga)
	if qty > 0 and harga > 0:
		return round(harga / qty, 4)
	return 0.0


def _default_opening_qty(uom, component_name):
	uom = (uom or "").upper()
	name = (component_name or "").upper()
	if uom == "GRAM":
		if "ICE" in name:
			return 100000
		return 10000
	if uom == "ML":
		return 50000
	if uom in ("PCS", "UNITS"):
		if "CUP" in name or "BOTOL" in name:
			return 500
		return 1000
	if uom == "SLICE":
		return 300
	if "BOTOL" in uom:
		return 200
	if uom == "SACHET":
		return 500
	return 100


def collect_components_from_menu(menu_path: Path) -> OrderedDict:
	items: OrderedDict[str, dict] = OrderedDict()
	if not menu_path.exists():
		return items

	with menu_path.open(encoding="utf-8") as handle:
		for row in csv.DictReader(handle):
			component = (row.get("Komponen") or "").strip()
			if not component:
				continue
			uom = (row.get("UOM") or "Nos").strip()
			rate = _unit_rate(row.get("Qty"), row.get("Harga"))
			current = items.get(component)
			if not current or (rate and rate > current["rate"]):
				items[component] = {
					"uom": uom,
					"rate": rate,
					"qty": _default_opening_qty(uom, component),
				}
	return items


def write_stock_csv(components: OrderedDict, output_path: Path) -> None:
	with output_path.open("w", newline="", encoding="utf-8") as handle:
		writer = csv.writer(handle)
		writer.writerow(STOCK_HEADERS)
		for name, data in components.items():
			writer.writerow([name, data["qty"], data["uom"], data["rate"] or ""])


def read_menu_rows(menu_path: Path) -> list[list]:
	rows = [MENU_HEADERS]
	with menu_path.open(encoding="utf-8") as handle:
		reader = csv.DictReader(handle)
		for row in reader:
			rows.append([row.get(header, "") for header in MENU_HEADERS])
	return rows


def write_workbook(menu_path: Path, components: OrderedDict, output_path: Path) -> None:
	from openpyxl import Workbook
	from openpyxl.styles import Font, PatternFill
	from openpyxl.utils import get_column_letter

	wb = Workbook()
	header_fill = PatternFill("solid", fgColor="FFF3CD")
	header_font = Font(bold=True)

	# Sheet Menu
	ws_menu = wb.active
	ws_menu.title = "Menu"
	for row in read_menu_rows(menu_path):
		ws_menu.append(row)
	for cell in ws_menu[1]:
		cell.font = header_font
		cell.fill = header_fill

	# Sheet Stok Awal
	ws_stock = wb.create_sheet("Stok Awal")
	ws_stock.append(STOCK_HEADERS)
	for name, data in components.items():
		ws_stock.append([name, data["qty"], data["uom"], data["rate"] or None])
	for cell in ws_stock[1]:
		cell.font = header_font
		cell.fill = PatternFill("solid", fgColor="D4EDDA")

	notes = ws_stock.cell(row=1, column=6, value="Catatan")
	notes.font = header_font
	notes.fill = header_fill
	row_idx = 2
	for name, data in components.items():
		note = ""
		if not data["rate"]:
			note = "Isi Harga satuan dari faktur pembelian"
		ws_stock.cell(row=row_idx, column=6, value=note)
		row_idx += 1

	for sheet in (ws_menu, ws_stock):
		for idx, column_cells in enumerate(sheet.columns, start=1):
			max_len = 0
			for cell in column_cells:
				if cell.value is not None:
					max_len = max(max_len, len(str(cell.value)))
			sheet.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 48)

	wb.save(output_path)


def main():
	components = collect_components_from_menu(MENU_CSV)
	if not components:
		raise SystemExit(f"No components found in {MENU_CSV}")

	write_stock_csv(components, STOCK_CSV)
	write_workbook(MENU_CSV, components, WORKBOOK_XLSX)

	missing_rate = sum(1 for data in components.values() if not data["rate"])
	print(f"Components: {len(components)}")
	print(f"Wrote {STOCK_CSV}")
	print(f"Wrote {WORKBOOK_XLSX} (sheets: Menu, Stok Awal)")
	if missing_rate:
		print(f"Warning: {missing_rate} items have empty Harga — fill before import.")


if __name__ == "__main__":
	main()
